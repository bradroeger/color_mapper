import json
import argparse
from pathlib import Path
from collections import defaultdict
from openpyxl import load_workbook
from openpyxl.styles.fills import PatternFill
from openpyxl.styles.colors import COLOR_INDEX


def load_config(config_path: Path) -> dict:
    """
    Load color→text mappings from JSON and normalize keys.
    Accepts keys in forms like:
      - "FFFF0000" (ARGB) or "FF0000" (RGB) or "#FF0000"
      - "rgb:FFFF0000" or "rgb:FF0000"
      - "indexed:64"
      - "theme:0"
    Stored internally as one of:
      - "rgb:FFFFFFFF"
      - "indexed:<int>"
      - "theme:<int>"
    """
    with open(config_path, "r", encoding="utf-8") as f:
        raw = json.load(f)

    def norm_key(k: str) -> str:
        s = str(k).strip().lower()
        # allow "indexed:n" and "theme:n" explicitly
        if s.startswith("indexed:"):
            n = s.split(":", 1)[1].strip()
            return f"indexed:{int(n)}"
        if s.startswith("theme:"):
            n = s.split(":", 1)[1].strip()
            return f"theme:{int(n)}"

        # allow "rgb:" prefix
        if s.startswith("rgb:"):
            s = s[4:].strip()

        # allow "#RRGGBB" or "RRGGBB" or "AARRGGBB"
        s = s.lstrip("#").upper()

        if len(s) == 6:      # RGB -> ARGB with FF alpha
            s = "FF" + s
        elif len(s) == 8:    # already ARGB
            pass
        else:
            raise ValueError(f"Color '{k}' must be 6 or 8 hex digits (or use theme:/indexed:).")

        return f"rgb:{s}"

    out = {}
    for k, v in raw.items():
        out[norm_key(k)] = v
    return out

def _color_to_key(c):
    if c is None:
        return None
    t = getattr(c, "type", None)
    if t == "rgb" and getattr(c, "rgb", None):
        rgb = c.rgb.upper()
        if len(rgb) == 6:          # RGB -> ARGB
            rgb = "FF" + rgb
        return f"rgb:{rgb}"
    if t == "indexed" and getattr(c, "indexed", None) is not None:
        return f"indexed:{int(c.indexed)}"
    if t == "theme" and getattr(c, "theme", None) is not None:
        return f"theme:{int(c.theme)}"
    return None


def color_key_from_fill(fill):
    """
    Extract a comparable color key from any openpyxl fill.
    Works with PatternFill, GradientFill, and StyleProxy-wrapped fills.
    """
    if fill is None:
        return None

    # Unwrap StyleProxy if present
    fill = getattr(fill, "_target", fill)

    # Some files set colors even if fill_type/patternType is None,
    # so don't early-return based on that—just try the color attrs.
    for attr in ("start_color", "fgColor", "end_color"):
        c = getattr(fill, attr, None)
        key = _color_to_key(c)
        if key:
            return key
    return None


def process(
    excel_path: Path,
    config_path: Path,
    output_path: Path | None,
    mode: str,
    sheet_name: str | None,
    debug: bool,
):
    color_map = load_config(config_path)

    wb = load_workbook(excel_path)
    ws = wb[sheet_name] if sheet_name else wb.active

    if mode not in {"row", "column"}:
        raise ValueError("mode must be 'row' or 'column'.")

    # For 'row' mode, write into a single new column at the far right (fixed).
    # Capture once so we don’t keep moving the goalpost as we write values.
    if mode == "row":
        target_col = ws.max_column + 1
        written_rows = set()

    # For 'column' mode, we’ll append below the last used cell in the column of each match.
    # Pre-calc last used row per column.
    if mode == "column":
        last_used = defaultdict(int)
        for col_idx in range(1, ws.max_column + 1):
            # scan from bottom up to find last used row for this column
            for r in range(ws.max_row, 0, -1):
                if ws.cell(row=r, column=col_idx).value not in (None, ""):
                    last_used[col_idx] = r
                    break
            if last_used[col_idx] == 0:
                last_used[col_idx] = 0  # empty column so far

    matches = 0
    for row in ws.iter_rows():
        for cell in row:
            print(cell)
            ck = color_key_from_fill(cell.fill)
            print(ck)
            if debug:
                # Only print something helpful; avoid spamming None lines
                if ck:
                    print(f"DEBUG: R{cell.row}C{cell.column}: {ck}")

            if ck and ck in color_map:
                out_text = color_map[ck]
                if mode == "row":
                    if cell.row not in written_rows:
                        ws.cell(row=cell.row, column=target_col, value=out_text)
                        written_rows.add(cell.row)
                        matches += 1
                else:  # column mode
                    col_idx = cell.column
                    last_used[col_idx] += 1
                    ws.cell(row=last_used[col_idx], column=col_idx, value=out_text)
                    matches += 1

    if output_path is None:
        output_path = excel_path.with_stem(excel_path.stem + "_output")

    wb.save(output_path)
    print(f"✅ Done. {matches} match(es) written. Saved to: {output_path}")


def main():
    p = argparse.ArgumentParser(
        description="Append notes based on Excel cell fill colors (openpyxl)."
    )
    p.add_argument("excel", help="Path to input .xlsx")
    p.add_argument("config", help="Path to config JSON with color→text pairs")
    p.add_argument("-o", "--output", help="Path to save .xlsx (defaults to *_output.xlsx)")
    p.add_argument(
        "--mode",
        choices=["row", "column"],
        default="row",  # closer to “append to that column”
        help="row = write once per matching row in a new right-most column;"
             " column = append below last used cell of the matched column (default)",
    )
    p.add_argument("--sheet", help="Optional: specific worksheet name")
    p.add_argument("--debug", action="store_true", help="Print debug info for detected colors")
    args = p.parse_args()

    process(
        excel_path=Path(args.excel),
        config_path=Path(args.config),
        output_path=Path(args.output) if args.output else None,
        mode=args.mode,
        sheet_name=args.sheet,
        debug=args.debug,
    )


if __name__ == "__main__":
    main()
